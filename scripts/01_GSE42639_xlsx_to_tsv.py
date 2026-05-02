#!/usr/bin/env python3
"""
GSE42639 xlsx → 3 个 tsv：
  intensity (probe × sample)、detection pval (probe × sample)、sample metadata
xlsx 三行表头：Row1=GSM, Row2=描述名, Row3=ID_REF + .AVG_Signal/.Detection Pval 列。
单一已知错字: GSM1047149 名中 Sd02 -> d02。
"""
import re
import sys
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path("/Users/wuxiuxiang/project/dongmei/OptiSyn")
SRC = ROOT / "raw_data/00_raw_data/GSE42639_non-normalized.xlsx"
OUT = ROOT / "processed_data/GSE42639"
OUT.mkdir(parents=True, exist_ok=True)

CELLS = ("Nh", "Am", "Ne", "Mo", "Ly")
TREATS = ("Sham", "TX91", "0.6LD50PR8", "10LD50PR8", "100LD50PR8")


def parse_name(name: str):
    n = name.replace("-", "_").replace("_Sd02_", "_d02_")
    cell = next((c for c in CELLS if f"_{c}_" in n + "_"), None)
    treat = next((t for t in TREATS if t in n), None)
    m_day = re.search(r"_d(\d+)", n)
    m_rep = re.search(r"rep(\d+\w?)", n)
    return {
        "cell": cell,
        "treat": treat,
        "day": "d" + m_day.group(1) if m_day else None,
        "rep": m_rep.group(1) if m_rep else None,
    }


def main():
    print(f"reading: {SRC}")
    wb = load_workbook(SRC, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    gsm_row = next(it)
    name_row = next(it)
    _ = next(it)  # field row, not needed (we know layout)

    # samples: list of (gsm, full_name, intensity_col_idx, detection_col_idx)
    samples = []
    for c in range(1, len(gsm_row), 2):
        gsm = gsm_row[c]
        name = name_row[c]
        if gsm is None:
            continue
        samples.append((gsm, name, c, c + 1))
    print(f"  samples: {len(samples)}")

    # write metadata
    meta_path = OUT / "GSE42639_sample_meta.tsv"
    with meta_path.open("w") as f:
        f.write("sample_id\tgsm\tfull_name\tcell\ttreat\tday\trep\n")
        for gsm, name, _, _ in samples:
            p = parse_name(name)
            sid = f"{gsm}_{p['cell']}_{p['treat']}_{p['rep']}"
            f.write(
                f"{sid}\t{gsm}\t{name}\t{p['cell']}\t{p['treat']}\t{p['day']}\t{p['rep']}\n"
            )
    print(f"  wrote: {meta_path.name}")

    # build sample_id list aligned with column order
    sample_ids = []
    for gsm, name, _, _ in samples:
        p = parse_name(name)
        sample_ids.append(f"{gsm}_{p['cell']}_{p['treat']}_{p['rep']}")

    # stream data rows → write intensity + detection in parallel
    int_path = OUT / "GSE42639_intensity.tsv"
    det_path = OUT / "GSE42639_detection.tsv"
    n_data = 0
    with int_path.open("w") as fi, det_path.open("w") as fd:
        header = "probe_id\t" + "\t".join(sample_ids) + "\n"
        fi.write(header)
        fd.write(header)
        for row in it:
            probe = row[0]
            if probe is None:
                continue
            ints = [row[s[2]] for s in samples]
            dets = [row[s[3]] for s in samples]
            fi.write(str(probe) + "\t" + "\t".join("" if v is None else str(v) for v in ints) + "\n")
            fd.write(str(probe) + "\t" + "\t".join("" if v is None else str(v) for v in dets) + "\n")
            n_data += 1
            if n_data % 10000 == 0:
                print(f"  ...{n_data} probes")
    print(f"  wrote: {int_path.name} + {det_path.name} ({n_data} probes)")


if __name__ == "__main__":
    main()
